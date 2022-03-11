from openpyxl import load_workbook
my_excel = 'Product-spec.xlsx'

#only_data gaves you data, not formula in excel
wb = load_workbook(my_excel, data_only=True)
 
sheet = wb.active

#iterete through rows
big_list = []
for i in range(1, sheet.max_row+1):
    inside_list = ()
    a = sheet[f'A{str(i)}'].value
    c = sheet[f'C{str(i)}'].value
    l = sheet[f'L{str(i)}'].value
    f = sheet[f'O{str(i)}'].value
    #inside tuple
    inside_list = (a,c,l,f)
    big_list.append(inside_list)
    #remove duplicates
    unique = list(set(big_list))
    if inside_list in unique:
        new_val = unique.index(inside_list)+1
        sheet[f'S{str(i)}'] = new_val

wb.save(my_excel)